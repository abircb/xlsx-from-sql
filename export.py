from openpyxl import Workbook
from optparse import OptionParser
import mysql.connector
import progressbar

# Enter DB details here
"""
hostS     #source database
userS     #username
passwordS #password'
databaseS
portS
"""

parser = OptionParser(conflict_handler="resolve")

parser.add_option("-h", "--host", dest="host", help="DB hostname", default=hostS)
parser.add_option("-u", "--user", dest="user", help="DB username", default=userS)
parser.add_option(
    "-p", "--password", dest="password", help="DB password", default=passwordS
)
parser.add_option(
    "-d", "--database", dest="database", help="Database name", default=databaseS
)
parser.add_option(
    "-o", "--output", dest="output", help="Output xlsx filename", default=""
)
parser.add_option(
    "-v",
    "--verbose",
    action="store_true",
    dest="verbose",
    default=True,
    help="Report progress [default]",
)
parser.add_option(
    "-q", "--quiet", action="store_false", dest="verbose", help="Be quiet"
)

(options, args) = parser.parse_args()

if not options.database:
    parser.error("Unknown Database name")

table = # Enter your output table's name
options.output = table + ".xlsx"


cnx = mysql.connector.connect(
    user=options.user,
    password=options.password,
    host=options.host,
    database=options.database,
)
cursor = cnx.cursor()

book = Workbook(write_only=True):
title = table[-31:]
sheet = book.create_sheet()
sheet.title = title

# count query to create progress bar
query_count = # "SELECT count(1) FROM TABLE_NAME WHERE PARAMETER = 'SOMETHING'"
cursor.execute(query_count)
(rows,) = cursor.fetchone()

# your query
query = # Enter your query here
cursor.execute(query)

field_names = [i[0] for i in cursor.description]
print(field_names)
sheet.append(field_names)

print("Exporting %s" % table)
if options.verbose:
    barCurrent = progressbar.ProgressBar(
        maxval=rows,
        widgets=[progressbar.Bar("=", "[", "]"), " ", progressbar.Percentage()],
    ).start()

i = 0
for row in cursor:
    row = [(x.decode("utf-8") if type(x) is bytearray else x) for x in row]
    sheet.append(row)
    i += 1
    if options.verbose:
        barCurrent.update(i)

if options.verbose:
    barCurrent.finish()

# adding your query to the Excel sheet
title = "Query"
sheet = book.create_sheet()
sheet.title = title
sheet.append(["Query"])
sheet.append([(query.decode("utf-8") if type(query) is bytearray else query)])

cursor.close()
if options.verbose:
    print("Writing output file...")

book.save(options.output)
cnx.close()
